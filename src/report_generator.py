"""
report_generator.py
Generates two Excel outputs:
  1. relatorio_critica_cadastral.xlsx  — actuarial report (4 sheets)
  2. powerbi_data.xlsx                 — flat tables for Power BI to consume
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import os

# Colours
C_HDR   = "1F3864"
C_SUB   = "2E75B6"
C_LIGHT = "DEEAF1"
C_CRIT  = "FF4C4C"
C_ALERT = "FFB347"
C_OK    = "C6EFCE"
C_WHITE = "FFFFFF"

def _b(): s = Side(style="thin", color="AAAAAA"); return Border(left=s,right=s,top=s,bottom=s)

def _hdr(ws, row, col, val, bg=C_HDR, size=11):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(color="FFFFFF", bold=True, size=size)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _b()
    return c

def _cell(ws, row, col, val, bg=C_WHITE, bold=False, fmt=None, align="left"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _b()
    if fmt: c.number_format = fmt
    return c

SEV_BG = {"CRITICO": C_CRIT, "ALERTA": C_ALERT}


# ── Sheet builders ────────────────────────────────────────────────────────────

def _sheet_summary(wb, summ, all_issues):
    ws = wb.active
    ws.title = "Sumário Executivo"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 22

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value     = "CRÍTICA DA BASE CADASTRAL — AVALIAÇÃO ATUARIAL"
    c.font      = Font(bold=True, size=15, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C_HDR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:B2")
    c2 = ws["A2"]
    c2.value     = (f"Data-base: {summ['data_base']}  |  "
                    f"Res. PREVIC 7/2022 + CPA 017/2019 IBA")
    c2.font      = Font(italic=True, size=9, color="FFFFFF")
    c2.fill      = PatternFill("solid", fgColor=C_SUB)
    c2.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    row = 4
    _hdr(ws, row, 1, "INDICADOR POPULACIONAL", bg=C_SUB)
    _hdr(ws, row, 2, "VALOR",                  bg=C_SUB)
    row += 1

    items = [
        ("Ativos",                    summ["n_ativos"],                  "#,##0"),
        ("Assistidos (Beneficiários)",summ["n_assistidos"],              "#,##0"),
        ("Diferidos (BPD)",           summ["n_diferidos"],               "#,##0"),
        ("TOTAL",                     summ["n_total"],                   "#,##0"),
        ("Razão Assistidos / Ativos", summ["razao_assistidos_ativos"],   "0.000"),
        ("Média Idade — Ativos",      summ["ativos_media_idade"],        "0.0"),
        ("Média Salário de Contrib.", summ["ativos_media_salario"],      "R$ #,##0.00"),
        ("Massa Salarial / mês",      summ["ativos_total_massa"],        "R$ #,##0.00"),
        ("Média Idade — Assistidos",  summ["assistidos_media_idade"],    "0.0"),
        ("Benefício Médio / mês",     summ["assistidos_media_beneficio"],"R$ #,##0.00"),
        ("Total Benefícios / mês",    summ["assistidos_total_beneficio"],"R$ #,##0.00"),
    ]
    for label, val, fmt in items:
        bg = C_LIGHT if row % 2 == 0 else C_WHITE
        bold = label == "TOTAL"
        _cell(ws, row, 1, label, bg=bg, bold=bold)
        _cell(ws, row, 2, val,   bg=bg, bold=bold, fmt=fmt, align="right")
        row += 1

    row += 1
    n_crit  = (all_issues["SEVERIDADE"] == "CRITICO").sum()
    n_alert = (all_issues["SEVERIDADE"] == "ALERTA").sum()

    _hdr(ws, row, 1, "RESULTADO DA CRÍTICA", bg=C_SUB)
    _hdr(ws, row, 2, "QTD",                  bg=C_SUB)
    row += 1

    for label, val, bg in [
        ("Inconsistências CRÍTICAS",   n_crit,          C_CRIT),
        ("Inconsistências ALERTAS",    n_alert,         C_ALERT),
        ("Total de ocorrências",       len(all_issues), C_LIGHT),
    ]:
        _cell(ws, row, 1, label, bg=bg, bold=True)
        _cell(ws, row, 2, val,   bg=bg, bold=True, fmt="#,##0", align="right")
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:B{row}")
    nota = ws.cell(row=row, column=1,
        value=("NOTA REGULATÓRIA: Registros com inconsistências CRÍTICAS não "
               "devem ser utilizados na avaliação atuarial sem correção prévia "
               "ou justificativa formal do atuário responsável. "
               "Ref: CPA 017/2019 IBA, seção 4.2."))
    nota.font = Font(italic=True, size=9)
    nota.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 50


def _sheet_issues(wb, all_issues):
    ws = wb.create_sheet("Inconsistências")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    cols = [("ID Participante",15),("Grupo",12),("Campo",20),
            ("Valor Atual",20),("Severidade",13),("Código",10),("Descrição",75)]
    for i, (title, width) in enumerate(cols, 1):
        _hdr(ws, 1, i, title)
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.merge_cells("A2:G2")
    leg = ws["A2"]
    leg.value = ("CRÍTICO = corrigir antes da avaliação  |  "
                 "ALERTA = atuário analisa e justifica  |  "
                 "Ref: Res. PREVIC 7/2022 Art. 8 + CPA 017/2019 IBA")
    leg.font      = Font(italic=True, size=9)
    leg.fill      = PatternFill("solid", fgColor=C_LIGHT)
    leg.alignment = Alignment(horizontal="center")

    sorted_df = all_issues.sort_values(["SEVERIDADE","CODIGO"])
    for r_i, (_, row) in enumerate(sorted_df.iterrows(), start=3):
        bg = SEV_BG.get(row["SEVERIDADE"], C_WHITE)
        bold = row["SEVERIDADE"] == "CRITICO"
        for c_i, f in enumerate(
            ["ID_PARTICIPANTE","GRUPO","CAMPO","VALOR_ATUAL",
             "SEVERIDADE","CODIGO","DESCRICAO"], start=1
        ):
            c = ws.cell(row=r_i, column=c_i, value=row.get(f,""))
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(bold=bold)
            c.alignment = Alignment(vertical="center", wrap_text=(c_i==7))
            c.border = _b()
        ws.row_dimensions[r_i].height = 16


def _sheet_clean(wb, df_a, all_issues):
    ws = wb.create_sheet("Base Limpa — Ativos")
    ws.sheet_view.showGridLines = False

    crit_ids = set(all_issues[all_issues["SEVERIDADE"]=="CRITICO"]["ID_PARTICIPANTE"])
    df_clean = df_a[~df_a["ID_PARTICIPANTE"].isin(crit_ids)].copy()

    cols = list(df_clean.columns)
    for i, col in enumerate(cols, 1):
        _hdr(ws, 1, i, col)
        ws.column_dimensions[get_column_letter(i)].width = 18

    for r_i, (_, row) in enumerate(df_clean.iterrows(), start=2):
        bg = C_LIGHT if r_i % 2 == 0 else C_WHITE
        for c_i, col in enumerate(cols, 1):
            val = row[col]
            c   = ws.cell(row=r_i, column=c_i,
                          value=None if pd.isna(val) else val)
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = _b()

    n_excl = len(df_a) - len(df_clean)
    ws.cell(row=len(df_clean)+3, column=1,
            value=(f"Base limpa: {len(df_clean):,} registros válidos de "
                   f"{len(df_a):,} originais. "
                   f"{n_excl} excluídos por inconsistências críticas.")).font = \
        Font(italic=True, bold=True)


def _sheet_by_type(wb, all_issues):
    ws = wb.create_sheet("Análise por Tipo")
    ws.sheet_view.showGridLines = False

    for i, (t, w) in enumerate([
        ("Código",12),("Severidade",13),("Campo",22),
        ("Ocorrências",14),("% do Total",12)
    ], 1):
        _hdr(ws, 1, i, t)
        ws.column_dimensions[get_column_letter(i)].width = w

    freq = (all_issues.groupby(["CODIGO","SEVERIDADE","CAMPO"])
            .size().reset_index(name="n")
            .sort_values("n", ascending=False))
    total = len(all_issues) or 1

    for r_i, (_, row) in enumerate(freq.iterrows(), start=2):
        bg = SEV_BG.get(row["SEVERIDADE"], C_WHITE)
        pct = row["n"] / total
        for c_i, (val, fmt) in enumerate([
            (row["CODIGO"],     None),
            (row["SEVERIDADE"], None),
            (row["CAMPO"],      None),
            (row["n"],          "#,##0"),
            (pct,               "0.0%"),
        ], 1):
            c = ws.cell(row=r_i, column=c_i, value=val)
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = _b()
            if fmt: c.number_format = fmt

    chart = BarChart()
    chart.type = "col"; chart.style = 10
    chart.title = "Inconsistências por Tipo"
    chart.y_axis.title = "Ocorrências"
    chart.width = 20; chart.height = 12
    data_ref = Reference(ws, min_col=4, min_row=1, max_row=min(len(freq)+1,15))
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=min(len(freq)+1,15))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "G2")


# ── Power BI flat export ──────────────────────────────────────────────────────

def build_powerbi_export(df_a, df_s, df_d, all_issues, summ,
                          output_path="../results/reports/powerbi_data.xlsx"):
    """
    Exports flat tables optimised for Power BI import:
      - pop_ativos, pop_assistidos, pop_diferidos  (population with age/service)
      - inconsistencias                             (all issues)
      - sumario                                     (KPI table, 1 row per metric)
      - freq_inconsistencias                        (for bar chart visual)
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    REF = pd.Timestamp("2024-12-31")

    def enrich_pop(df, situacao):
        df = df.copy()
        df["SITUACAO_GRUPO"] = situacao
        if "DT_NASCIMENTO" in df.columns:
            dt = pd.to_datetime(df["DT_NASCIMENTO"], dayfirst=True, errors="coerce")
            df["IDADE"] = ((REF - dt).dt.days / 365.25).round(1)
            df["FAIXA_ETARIA"] = pd.cut(
                df["IDADE"],
                bins=[0,25,35,45,55,65,75,120],
                labels=["< 25","25–34","35–44","45–54","55–64","65–74","≥ 75"],
                right=False
            ).astype(str)
        if "DT_ADMISSAO_PLANO" in df.columns:
            dt2 = pd.to_datetime(df["DT_ADMISSAO_PLANO"], dayfirst=True, errors="coerce")
            df["TEMPO_PLANO_ANOS"] = ((REF - dt2).dt.days / 365.25).round(1)
        return df

    df_a_enr = enrich_pop(df_a, "ATIVO")
    df_s_enr = enrich_pop(df_s, "ASSISTIDO")
    df_d_enr = enrich_pop(df_d, "DIFERIDO")

    # Freq table for Power BI bar chart
    freq = (all_issues.groupby(["CODIGO","SEVERIDADE","CAMPO"])
            .size().reset_index(name="OCORRENCIAS"))
    freq["PCT_TOTAL"] = (freq["OCORRENCIAS"] / len(all_issues) * 100).round(1)

    # KPI summary as flat table (1 row per metric — Power BI card visuals)
    kpi = pd.DataFrame([
        {"INDICADOR": k, "VALOR": v} for k, v in summ.items()
    ])

    with pd.ExcelWriter(output_path, engine="openpyxl") as w:
        df_a_enr.to_excel(w, sheet_name="POP_ATIVOS",        index=False)
        df_s_enr.to_excel(w, sheet_name="POP_ASSISTIDOS",    index=False)
        df_d_enr.to_excel(w, sheet_name="POP_DIFERIDOS",     index=False)
        all_issues.to_excel(w, sheet_name="INCONSISTENCIAS", index=False)
        freq.to_excel(w, sheet_name="FREQ_INCONSISTENCIAS",  index=False)
        kpi.to_excel(w, sheet_name="SUMARIO_KPI",            index=False)

    print(f"Power BI data: {output_path}")
    return output_path


# ── Main actuarial report ─────────────────────────────────────────────────────

def build_report(df_a, df_s, df_d, iss_a, iss_s, iss_d, summ,
                 output_path="../results/reports/relatorio_critica_cadastral.xlsx"):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    all_issues = pd.concat([iss_a, iss_s, iss_d], ignore_index=True)

    wb = Workbook()
    _sheet_summary(wb, summ, all_issues)
    _sheet_issues(wb, all_issues)
    _sheet_clean(wb, df_a, all_issues)
    _sheet_by_type(wb, all_issues)

    wb.save(output_path)
    print(f"Relatório atuarial: {output_path}")

    # Also build Power BI export
    pbi_path = output_path.replace(
        "relatorio_critica_cadastral.xlsx", "powerbi_data.xlsx")
    build_powerbi_export(df_a, df_s, df_d, all_issues, summ, pbi_path)

    return output_path
